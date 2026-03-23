#!/usr/bin/env python3
"""
Good Life Bahamas — Weekly Property Page CRO Orchestration Script
=================================================================
Runs every week to:
  1. Pull GA4 data filtered to /property/* paths
  2. Pull HubSpot pipeline data (shared with landing page cycle)
  3. Identify top and bottom performing property pages by inquiry rate
  4. Derive a hypothesis for property page improvements
  5. Generate property-pages/week-N.html — a CRO briefing document
  6. Append a metrics row to GoodLife_Property_CRO_Tracker.xlsx
  7. Update property-pages/index.html with the new week card
  8. Commit and push

Usage:
  python3 cro-analytics/weekly_cycle_property.py   # from repo root
  python3 weekly_cycle_property.py                 # from cro-analytics/

Required env vars (same .env as weekly_cycle.py):
  GA4_KEY_FILE     — path to service account JSON
  GA4_PROPERTY_ID  — GA4 property ID (e.g. 375125067)
  HUBSPOT_TOKEN    — HubSpot private app access token
"""

import os
import sys
import re
import subprocess
from datetime import datetime, timedelta
from pathlib import Path

# ── Load .env if present ──────────────────────────────────────────────────────
try:
    from dotenv import load_dotenv
    _env_path = Path(__file__).parent / '.env'
    if _env_path.exists():
        load_dotenv(_env_path)
        print(f"[env] Loaded .env from {_env_path}")
except ImportError:
    pass

# ── Paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).parent
REPO_ROOT    = SCRIPT_DIR.parent
PROPERTY_DIR = REPO_ROOT / 'property-pages'
TRACKER_PATH = REPO_ROOT / 'GoodLife_Property_CRO_Tracker.xlsx'
INDEX_PATH   = PROPERTY_DIR / 'index.html'

# ── Import shared infrastructure from cro_report ─────────────────────────────
sys.path.insert(0, str(SCRIPT_DIR))
from cro_report import (
    pull_hubspot_data,
    pull_clarity_data,
    clarity_insights,
    get_ga4_client,
    ga4_report,
)

MIN_SESSIONS = 10  # minimum sessions to consider a property for ranking


# ── GA4: property page data pull ─────────────────────────────────────────────

def pull_property_ga4_data(days=7):
    """
    Pull GA4 metrics for all /property/* pages.

    Returns:
      property_pages   — dict keyed by slug, with sessions/bounce/duration/inquiries
      total_sessions   — int: sum of sessions across all property pages
      total_inquiries  — int: sum of inquiry events across all property pages
      overall_inquiry_rate — float: total_inquiries / total_sessions * 100
      by_device        — list of device rows (shared with landing page cycle)
    """
    client = get_ga4_client()
    start  = f'{days}daysAgo'
    end    = 'today'
    results = {}

    # 1. Per-page sessions + bounce + duration
    rows = ga4_report(
        client,
        dimensions=['pagePath'],
        metrics=['sessions', 'bounceRate', 'averageSessionDuration', 'screenPageViews'],
        start_date=start, end_date=end,
        limit=500,
    )

    property_pages = {}
    for r in rows:
        path = r[0]
        if '/property/' not in path:
            continue
        # Normalise slug: /property/surf-house/ → surf-house
        slug = path.strip('/').split('property/')[-1].strip('/')
        if not slug:
            slug = 'unknown'
        property_pages[slug] = {
            'path':         path,
            'sessions':     int(r[1]),
            'bounce_rate':  float(r[2]),
            'avg_duration': float(r[3]),
            'pageviews':    int(r[4]),
            'inquiries':    0,
        }

    # 2. Inquiry events fired on property pages
    INQUIRY_EVENTS = {
        'form_submit', 'inquire_form_submit',
        'booknow_button_click_event', 'booking_form_button_click_event',
        'generate_lead', 'contact_form_submit',
    }
    rows = ga4_report(
        client,
        dimensions=['pagePath', 'eventName'],
        metrics=['eventCount'],
        start_date=start, end_date=end,
        limit=1000,
    )
    for r in rows:
        path, event_name, count = r[0], r[1], int(r[2])
        if '/property/' not in path or event_name not in INQUIRY_EVENTS:
            continue
        slug = path.strip('/').split('property/')[-1].strip('/')
        if not slug:
            slug = 'unknown'
        if slug in property_pages:
            property_pages[slug]['inquiries'] += count
        else:
            # Event fired on a path that didn't show in session data — add it
            property_pages[slug] = {
                'path': path, 'sessions': 0, 'bounce_rate': 0.0,
                'avg_duration': 0.0, 'pageviews': 0, 'inquiries': count,
            }

    # 3. Compute inquiry rate per property
    for data in property_pages.values():
        s = data['sessions']
        data['inquiry_rate'] = round(data['inquiries'] / max(s, 1) * 100, 3)

    # 4. Totals
    total_sessions  = sum(d['sessions']  for d in property_pages.values())
    total_inquiries = sum(d['inquiries'] for d in property_pages.values())
    results['property_pages']       = property_pages
    results['total_sessions']       = total_sessions
    results['total_inquiries']      = total_inquiries
    results['overall_inquiry_rate'] = round(total_inquiries / max(total_sessions, 1) * 100, 3)

    # 5. Device breakdown (site-wide, for mobile context)
    rows = ga4_report(
        client,
        dimensions=['deviceCategory'],
        metrics=['sessions', 'bounceRate'],
        start_date=start, end_date=end,
    )
    results['by_device'] = [
        {'device': r[0], 'sessions': int(r[1]), 'bounce_rate': float(r[2])}
        for r in rows
    ]

    return results


# ── Week detection ────────────────────────────────────────────────────────────

def detect_week_number() -> int:
    existing = list(PROPERTY_DIR.glob('week-*.html'))
    nums = []
    for f in existing:
        m = re.match(r'^week-(\d+)\.html$', f.name)
        if m:
            nums.append(int(m.group(1)))
    return (max(nums) + 1) if nums else 1


def get_week_date_range(week_number: int):
    end   = datetime.now()
    start = end - timedelta(days=7)
    offset = timedelta(days=7 * (week_number - 1))
    return (start - offset).strftime('%Y-%m-%d'), (end - offset).strftime('%Y-%m-%d')


# ── Property ranking ──────────────────────────────────────────────────────────

def rank_properties(ga4p: dict):
    """
    Returns (top_list, bottom_list) each being a list of (slug, data) tuples,
    sorted by inquiry_rate descending for top and ascending for bottom.
    Only considers properties with >= MIN_SESSIONS sessions.
    """
    pages = {
        slug: data
        for slug, data in ga4p.get('property_pages', {}).items()
        if data['sessions'] >= MIN_SESSIONS
    }
    if not pages:
        return [], []

    ranked = sorted(pages.items(), key=lambda x: -x[1]['inquiry_rate'])
    top    = ranked[:5]
    bottom = list(reversed(ranked))[:5]
    return top, bottom


def pick_top_property(ga4p: dict):
    """Returns (slug, data) for the best-performing property, or None."""
    top, _ = rank_properties(ga4p)
    return top[0] if top else None


# ── Hypothesis ────────────────────────────────────────────────────────────────

def pick_property_hypothesis(ga4p: dict, hs: dict, top: list, bottom: list) -> str:
    """
    Property-page specific decision tree for the weekly hypothesis.
    """
    overall_rate = ga4p.get('overall_inquiry_rate', 0)
    total_sess   = ga4p.get('total_sessions', 0)
    by_device    = {d['device']: d for d in ga4p.get('by_device', [])}
    mobile       = by_device.get('mobile', {})
    total_site   = sum(d['sessions'] for d in ga4p.get('by_device', []))
    mob_pct      = mobile.get('sessions', 0) / max(total_site, 1) * 100
    mob_bounce   = mobile.get('bounce_rate', 0) * 100

    funnel         = hs.get('funnel_summary', {})
    close_lost_pct = funnel.get('close_lost_pct', 0)

    # Avg bounce across property pages
    pages = ga4p.get('property_pages', {})
    bounces = [d['bounce_rate'] for d in pages.values() if d['sessions'] >= MIN_SESSIONS]
    avg_bounce = sum(bounces) / len(bounces) * 100 if bounces else 0

    # Avg duration
    durations = [d['avg_duration'] for d in pages.values() if d['sessions'] >= MIN_SESSIONS]
    avg_duration = sum(durations) / len(durations) if durations else 0

    top_slug  = top[0][0]  if top  else 'n/a'
    top_rate  = top[0][1]['inquiry_rate']  if top  else 0
    worst_slug = bottom[0][0] if bottom else 'n/a'

    # Decision tree
    if overall_rate < 1.0 and total_sess >= 50:
        return (
            f"Overall property page inquiry rate is critically low ({overall_rate:.2f}%). "
            f"Hypothesis: The primary CTA ('Inquire' / 'Book Now') is not visible above the fold "
            f"on most property pages. Adding a sticky inquiry button on mobile and moving the "
            f"pricing + availability section above the gallery will lift inquiry rate by 30%+."
        )

    if mob_pct > 65 and mob_bounce > 65:
        return (
            f"{mob_pct:.0f}% of traffic is mobile with a {mob_bounce:.1f}% bounce rate. "
            f"Hypothesis: Property page images are loading slowly on mobile and the "
            f"inquiry form is buried below the fold. Implementing lazy-loaded WebP images "
            f"and a tap-to-inquire sticky button will reduce mobile bounce by 20%+."
        )

    if avg_bounce > 70:
        return (
            f"Average property page bounce rate is {avg_bounce:.1f}%. "
            f"Hypothesis: Visitors land on a property page but the headline and hero image "
            f"don't match their search intent. Adding a clear value headline "
            f"('3BR Beachfront — Sleeps 8, from $X/night') above the gallery will anchor "
            f"visitors and reduce bounce by 15%+."
        )

    if avg_duration < 60 and total_sess >= 50:
        return (
            f"Average time on property pages is only {avg_duration:.0f}s. "
            f"Hypothesis: Visitors are not engaging with the content — likely the gallery "
            f"is below the fold and key details (beds, baths, highlights) are hidden in "
            f"a collapsed section. Front-loading a '5 reasons to book' highlights block "
            f"will increase avg session time and inquiry rate."
        )

    if top and top_rate > overall_rate * 2:
        return (
            f"'{top_slug}' converts at {top_rate:.2f}% vs site average {overall_rate:.2f}%. "
            f"Hypothesis: The '{top_slug}' page layout pattern (likely: prominent gallery, "
            f"visible pricing, and inline social proof) should be applied to "
            f"'{worst_slug}' — the lowest-converting property — to lift its inquiry rate "
            f"to at least {overall_rate:.2f}%."
        )

    if close_lost_pct > 70:
        return (
            f"{close_lost_pct:.1f}% of pipeline deals end in Close Lost. "
            f"Hypothesis: Property pages lack trust signals at the decision moment. "
            f"Adding a 'Host responds within 2 hours' badge, 3 recent guest reviews, "
            f"and a transparent cancellation policy summary next to the inquiry form "
            f"will improve lead-to-booking conversion."
        )

    return (
        f"Property pages are performing at {overall_rate:.2f}% overall inquiry rate "
        f"({total_sess:,} sessions). "
        f"Iterative hypothesis: Standardise the page template around the "
        f"'{top_slug}' pattern — lead with hero + availability, follow with highlights, "
        f"keep the inquiry form pinned on scroll — to create a consistent high-converting "
        f"experience across all {len(pages)} tracked properties."
    )


# ── Briefing HTML generation ──────────────────────────────────────────────────

BRIEFING_CSS = """
  *, *::before, *::after { margin: 0; padding: 0; box-sizing: border-box; }
  body { font-family: 'Mulish', system-ui, sans-serif; background: #0D1B2A;
         color: #E2E8F0; min-height: 100vh; -webkit-font-smoothing: antialiased; }
  a { text-decoration: none; color: inherit; }
  :root {
    --teal: #31B8C6; --teal-dark: #2AA3B0; --teal-glow: rgba(49,184,198,0.15);
    --orange: #F28F52; --dark: #0D1B2A; --darker: #080F18;
    --card-bg: #1A2B3C; --card-bg-2: #152232; --border: rgba(255,255,255,0.07);
    --border-teal: rgba(49,184,198,0.25); --green: #22C55E; --yellow: #EAB308;
    --red: #EF4444; --text-muted: rgba(255,255,255,0.45); --text-mid: rgba(255,255,255,0.65);
  }
  .wrap { max-width: 960px; margin: 0 auto; padding: 0 24px 80px; }
  .page-header { background: linear-gradient(180deg,#060D14 0%,#0D1B2A 100%);
    border-bottom: 1px solid var(--border); padding: 48px 24px 40px; text-align: center; }
  .eyebrow { display: inline-flex; align-items: center; gap: 8px;
    background: rgba(49,184,198,0.1); border: 1px solid var(--border-teal);
    color: var(--teal); padding: 5px 16px; border-radius: 50px;
    font-size: 0.72rem; font-weight: 800; text-transform: uppercase;
    letter-spacing: 0.1em; margin-bottom: 18px; }
  .page-header h1 { font-size: clamp(1.8rem,4vw,2.8rem); font-weight: 900;
    color: #fff; margin-bottom: 10px; line-height: 1.15; }
  .page-header h1 .teal { color: var(--teal); }
  .page-header .sub { color: var(--text-muted); font-size: 0.95rem;
    max-width: 580px; margin: 0 auto; line-height: 1.7; }
  .section { padding: 48px 0 0; }
  .section-label { font-size: 0.72rem; font-weight: 800; text-transform: uppercase;
    letter-spacing: 0.12em; color: var(--teal); margin-bottom: 18px;
    display: flex; align-items: center; gap: 8px; }
  .section-label .bar { width: 22px; height: 2px; background: var(--teal); border-radius: 2px; }
  .section h2 { font-size: 1.4rem; font-weight: 900; color: #fff; margin-bottom: 8px; }
  .card { background: var(--card-bg); border: 1px solid var(--border-teal);
    border-radius: 16px; padding: 24px 28px; margin-bottom: 16px; }
  .stats-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
    gap: 1px; background: var(--border-teal); border: 1px solid var(--border-teal);
    border-radius: 16px; overflow: hidden; margin-bottom: 32px; }
  .stat-cell { background: var(--card-bg-2); padding: 18px 20px; }
  .stat-val { font-size: 1.4rem; font-weight: 900; color: #fff; line-height: 1.2; }
  .stat-val.teal { color: var(--teal); }
  .stat-val.orange { color: var(--orange); }
  .stat-lbl { font-size: 0.75rem; color: var(--text-muted); font-weight: 600; margin-top: 3px; }
  .prop-table { width: 100%; border-collapse: collapse; font-size: 0.87rem; }
  .prop-table th { text-align: left; padding: 10px 12px; font-size: 0.72rem;
    font-weight: 700; color: var(--text-muted); text-transform: uppercase;
    letter-spacing: 0.08em; border-bottom: 1px solid var(--border); }
  .prop-table td { padding: 11px 12px; border-bottom: 1px solid var(--border);
    color: var(--text-mid); vertical-align: middle; }
  .prop-table tr:last-child td { border-bottom: none; }
  .prop-table .slug { color: #fff; font-weight: 700; }
  .badge { display: inline-block; padding: 3px 10px; border-radius: 50px;
    font-size: 0.72rem; font-weight: 800; }
  .badge-green { background: rgba(34,197,94,0.12); color: var(--green); border: 1px solid rgba(34,197,94,0.25); }
  .badge-red   { background: rgba(239,68,68,0.12);  color: var(--red);   border: 1px solid rgba(239,68,68,0.25); }
  .badge-teal  { background: rgba(49,184,198,0.12); color: var(--teal);  border: 1px solid var(--border-teal); }
  .hyp-box { background: rgba(49,184,198,0.06); border: 1px solid var(--border-teal);
    border-radius: 12px; padding: 20px 24px; }
  .hyp-box p { color: #CBD5E1; line-height: 1.75; font-size: 0.95rem; }
  .action-list { list-style: none; padding: 0; }
  .action-list li { display: flex; align-items: flex-start; gap: 10px;
    padding: 11px 0; border-bottom: 1px solid var(--border); font-size: 0.9rem;
    color: var(--text-mid); line-height: 1.6; }
  .action-list li:last-child { border-bottom: none; }
  .action-list .num { flex-shrink: 0; width: 24px; height: 24px; border-radius: 50%;
    background: rgba(49,184,198,0.15); color: var(--teal); font-weight: 800;
    font-size: 0.78rem; display: flex; align-items: center; justify-content: center; }
  .footer { margin-top: 56px; padding-top: 24px; border-top: 1px solid var(--border);
    text-align: center; color: var(--text-muted); font-size: 0.8rem; }
"""


def _rate_badge(rate: float, avg: float) -> str:
    if rate >= avg * 1.5:
        return f'<span class="badge badge-green">{rate:.2f}%</span>'
    elif rate < avg * 0.5:
        return f'<span class="badge badge-red">{rate:.2f}%</span>'
    return f'<span class="badge badge-teal">{rate:.2f}%</span>'


def _prop_rows(items: list, avg_rate: float, label: str) -> str:
    if not items:
        return f'<tr><td colspan="5" style="color:var(--text-muted);font-style:italic">No {label} data (need ≥{MIN_SESSIONS} sessions)</td></tr>'
    rows = ''
    for slug, d in items:
        sessions  = d['sessions']
        inquiries = d['inquiries']
        bounce    = round(d['bounce_rate'] * 100, 1)
        dur       = round(d['avg_duration'])
        rate_html = _rate_badge(d['inquiry_rate'], avg_rate)
        rows += (
            f'<tr>'
            f'<td class="slug">{slug}</td>'
            f'<td>{sessions:,}</td>'
            f'<td>{inquiries}</td>'
            f'<td>{bounce}%</td>'
            f'<td>{rate_html}</td>'
            f'</tr>'
        )
    return rows


def _action_items(hypothesis: str, top: list, bottom: list, ga4p: dict) -> str:
    """Generate 4-5 concrete action items based on data."""
    actions = []
    avg_rate = ga4p.get('overall_inquiry_rate', 0)

    if bottom:
        worst_slug = bottom[0][0]
        worst_rate = bottom[0][1]['inquiry_rate']
        actions.append(
            f"<strong>Priority fix — {worst_slug}:</strong> Inquiry rate is {worst_rate:.2f}% "
            f"vs {avg_rate:.2f}% site average. Check: Is the CTA button visible above the fold? "
            f"Is the inquiry form loading correctly on mobile?"
        )

    if top and bottom:
        top_slug = top[0][0]
        actions.append(
            f"<strong>Apply winner pattern:</strong> Study '{top_slug}' — the top-converting property. "
            f"Document its layout (hero, highlights order, CTA placement) and apply the same "
            f"structure to the bottom 3 properties."
        )

    by_device = {d['device']: d for d in ga4p.get('by_device', [])}
    mobile = by_device.get('mobile', {})
    total_site = sum(d['sessions'] for d in ga4p.get('by_device', []))
    mob_pct = mobile.get('sessions', 0) / max(total_site, 1) * 100
    if mob_pct > 55:
        actions.append(
            f"<strong>Mobile check ({mob_pct:.0f}% of traffic):</strong> On a real mobile device, "
            f"visit the 3 lowest-converting properties. Time how long it takes to find and fill "
            f"the inquiry form. Target: form accessible within 2 taps, loading in &lt;3 seconds."
        )

    actions.append(
        "<strong>Clarity review (15 min):</strong> Open the Clarity dashboard filtered to "
        "this week's date range. Watch 3 sessions on property pages that bounced within 30 seconds. "
        "Note the pattern — is it gallery load time, pricing not visible, or form abandonment?"
    )

    actions.append(
        "<strong>HubSpot tagging:</strong> For new deals this week, check whether they came "
        "through a property page (source = property page referral). If HubSpot contact records "
        "don't show this, add a 'Lead Source Page' custom property to improve attribution."
    )

    items_html = ''
    for i, text in enumerate(actions, 1):
        items_html += f'<li><span class="num">{i}</span><span>{text}</span></li>'
    return items_html


def generate_briefing_html(
    week_num: int,
    week_start: str,
    week_end: str,
    hypothesis: str,
    ga4p: dict,
    hs: dict,
    top: list,
    bottom: list,
) -> str:
    """Generate the week-N.html CRO briefing document for property pages."""
    now_str      = datetime.now().strftime('%Y-%m-%d')
    total_sess   = ga4p.get('total_sessions', 0)
    total_inq    = ga4p.get('total_inquiries', 0)
    overall_rate = ga4p.get('overall_inquiry_rate', 0)
    all_pages    = ga4p.get('property_pages', {})
    num_props    = len([d for d in all_pages.values() if d['sessions'] >= MIN_SESSIONS])

    by_device  = {d['device']: d for d in ga4p.get('by_device', [])}
    mobile     = by_device.get('mobile', {})
    total_site = sum(d['sessions'] for d in ga4p.get('by_device', []))
    mob_pct    = round(mobile.get('sessions', 0) / max(total_site, 1) * 100, 1)

    funnel       = hs.get('funnel_summary', {})
    new_deals    = hs.get('new_deals_period', 0)
    converted    = hs.get('converted_period', 0)
    booking_rate = round(converted / max(new_deals, 1) * 100, 1)

    top_slug  = top[0][0]  if top  else '—'
    top_rate  = top[0][1]['inquiry_rate'] if top else 0
    worst_slug = bottom[0][0] if bottom else '—'

    top_rows    = _prop_rows(top,    overall_rate, 'top-property')
    bottom_rows = _prop_rows(bottom, overall_rate, 'bottom-property')
    action_items = _action_items(hypothesis, top, bottom, ga4p)
    hyp_escaped = hypothesis.replace('<', '&lt;').replace('>', '&gt;')

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Property CRO — Week {week_num} — The Good Life Bahamas</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Mulish:wght@300;400;500;600;700;800;900&display=swap" rel="stylesheet">
<style>{BRIEFING_CSS}</style>
</head>
<body>

<header class="page-header">
  <div class="eyebrow">&#x1F3E0; Property Page CRO &mdash; Week {week_num}</div>
  <h1>Property Page <span class="teal">Optimization Briefing</span></h1>
  <p class="sub">Data window: {week_start} &rarr; {week_end} &nbsp;&bull;&nbsp; Generated {now_str}<br>
  {num_props} properties tracked &bull; Automated weekly cycle</p>
</header>

<div class="wrap">

  <!-- ── Summary stats ── -->
  <div class="section">
    <div class="section-label"><span class="bar"></span>This Week at a Glance</div>
    <div class="stats-grid">
      <div class="stat-cell">
        <div class="stat-val teal">{total_sess:,}</div>
        <div class="stat-lbl">Property Page Sessions</div>
      </div>
      <div class="stat-cell">
        <div class="stat-val">{total_inq}</div>
        <div class="stat-lbl">Inquiries from Property Pages</div>
      </div>
      <div class="stat-cell">
        <div class="stat-val orange">{overall_rate:.2f}%</div>
        <div class="stat-lbl">Overall Inquiry Rate</div>
      </div>
      <div class="stat-cell">
        <div class="stat-val">{mob_pct}%</div>
        <div class="stat-lbl">Mobile Traffic (site-wide)</div>
      </div>
      <div class="stat-cell">
        <div class="stat-val">{new_deals}</div>
        <div class="stat-lbl">New HubSpot Deals (7d)</div>
      </div>
      <div class="stat-cell">
        <div class="stat-val">{booking_rate}%</div>
        <div class="stat-lbl">Deal Booking Rate</div>
      </div>
    </div>
  </div>

  <!-- ── Hypothesis ── -->
  <div class="section">
    <div class="section-label"><span class="bar"></span>Week {week_num} Hypothesis</div>
    <div class="hyp-box">
      <p>{hyp_escaped}</p>
    </div>
  </div>

  <!-- ── Top performers ── -->
  <div class="section">
    <div class="section-label"><span class="bar"></span>Top Converting Properties</div>
    <div class="card" style="padding: 0; overflow: hidden;">
      <table class="prop-table">
        <thead>
          <tr>
            <th>Property</th>
            <th>Sessions</th>
            <th>Inquiries</th>
            <th>Bounce Rate</th>
            <th>Inquiry Rate</th>
          </tr>
        </thead>
        <tbody>{top_rows}</tbody>
      </table>
    </div>
  </div>

  <!-- ── Bottom performers ── -->
  <div class="section">
    <div class="section-label"><span class="bar"></span>Properties Needing Attention</div>
    <div class="card" style="padding: 0; overflow: hidden;">
      <table class="prop-table">
        <thead>
          <tr>
            <th>Property</th>
            <th>Sessions</th>
            <th>Inquiries</th>
            <th>Bounce Rate</th>
            <th>Inquiry Rate</th>
          </tr>
        </thead>
        <tbody>{bottom_rows}</tbody>
      </table>
    </div>
  </div>

  <!-- ── Action items ── -->
  <div class="section">
    <div class="section-label"><span class="bar"></span>This Week&rsquo;s Action Items</div>
    <div class="card">
      <ul class="action-list">{action_items}</ul>
    </div>
  </div>

  <div class="footer">
    Generated by weekly_cycle_property.py &bull; Good Life Bahamas CRO Programme &bull; Week {week_num}
  </div>

</div>
</body>
</html>"""


# ── Excel tracker ─────────────────────────────────────────────────────────────

TRACKER_COLUMNS = [
    'Week', 'Date Generated', 'Week Start', 'Week End',
    'Top Property', 'Top Property Inquiry Rate %',
    'Overall Inquiry Rate %', 'Total Property Sessions',
    'Total Inquiries', 'Mobile % of Sessions',
    'Num Properties Tracked', 'Avg Bounce Rate %',
    'Hypothesis (short)',
    'New HubSpot Deals (7d)', 'Converted (Booked, 7d)', 'Period Conversion %',
    'Close Lost % (all-time)', 'Booked or Beyond (all-time)',
    'Clarity Rage Click %', 'Clarity Dead Click %',
    'Clarity Quick-Back %', 'Clarity Avg Scroll Depth %',
    'Clarity Avg Active Time (sec)', 'Clarity Notes (manual)',
    'Properties Improved (manual)', 'HTML File', 'Clarity Dashboard Link',
]


def append_tracker_row(week_num, week_start, week_end, hypothesis,
                       ga4p, hs, clarity, top, html_filename):
    try:
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("[warn] openpyxl not installed — skipping Excel update. pip install openpyxl")
        return

    total_sess  = ga4p.get('total_sessions', 0)
    total_inq   = ga4p.get('total_inquiries', 0)
    overall_rate = ga4p.get('overall_inquiry_rate', 0)
    by_device   = {d['device']: d for d in ga4p.get('by_device', [])}
    mobile      = by_device.get('mobile', {})
    total_site  = sum(d['sessions'] for d in ga4p.get('by_device', []))
    mob_pct     = round(mobile.get('sessions', 0) / max(total_site, 1) * 100, 1)

    all_pages   = ga4p.get('property_pages', {})
    num_props   = len([d for d in all_pages.values() if d['sessions'] >= MIN_SESSIONS])
    bounces     = [d['bounce_rate'] for d in all_pages.values() if d['sessions'] >= MIN_SESSIONS]
    avg_bounce  = round(sum(bounces) / len(bounces) * 100, 1) if bounces else 0.0

    funnel       = hs.get('funnel_summary', {})
    new_deals    = hs.get('new_deals_period', 0)
    converted    = hs.get('converted_period', 0)
    period_conv  = round(converted / new_deals * 100, 1) if new_deals > 0 else 0.0
    close_lost   = funnel.get('close_lost_pct', 0)
    booked       = funnel.get('booked_or_beyond', 0)

    c_overall   = clarity.get('overall', {}) if clarity.get('available') else {}
    c_link      = clarity.get('clarity_link', '')
    hyp_short   = hypothesis[:120] + ('...' if len(hypothesis) > 120 else '')

    top_slug = top[0][0] if top else ''
    top_rate = top[0][1]['inquiry_rate'] if top else 0.0

    row_data = [
        week_num, datetime.now().strftime('%Y-%m-%d'), week_start, week_end,
        top_slug, top_rate, overall_rate, total_sess, total_inq, mob_pct,
        num_props, avg_bounce, hyp_short,
        new_deals, converted, period_conv, close_lost, booked,
        c_overall.get('rage_click_pct', ''), c_overall.get('dead_click_pct', ''),
        c_overall.get('quick_back_pct', ''), c_overall.get('avg_scroll_depth', ''),
        c_overall.get('avg_active_time_sec', ''), '',  # Clarity Notes (manual)
        '',  # Properties Improved (manual)
        html_filename, c_link,
    ]

    if TRACKER_PATH.exists():
        wb = load_workbook(TRACKER_PATH)
        ws = wb.active
    else:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = Workbook()
        ws = wb.active
        ws.title = 'Property CRO Tracker'

        hfill  = PatternFill(start_color='0D1B2A', end_color='0D1B2A', fill_type='solid')
        hfont  = Font(bold=True, color='31B8C6', size=10)
        hbord  = Border(bottom=Side(style='medium', color='31B8C6'))
        halign = Alignment(horizontal='center', wrap_text=True)

        for i, col_name in enumerate(TRACKER_COLUMNS, 1):
            cell = ws.cell(row=1, column=i, value=col_name)
            cell.font = hfont; cell.fill = hfill
            cell.border = hbord; cell.alignment = halign

        ws.freeze_panes = 'A2'

    next_row = ws.max_row + 1
    dfont = Font(size=9)
    alt   = PatternFill(start_color='1A2B3C', end_color='1A2B3C', fill_type='solid')
    rfill = alt if next_row % 2 == 0 else None

    for i, val in enumerate(row_data, 1):
        cell = ws.cell(row=next_row, column=i, value=val)
        cell.font = dfont
        cell.alignment = Alignment(vertical='top', wrap_text=(i == 13))
        if rfill:
            cell.fill = rfill

    wb.save(TRACKER_PATH)
    print(f"[tracker] Appended Week {week_num} row to {TRACKER_PATH}")


# ── property-pages/index.html update ─────────────────────────────────────────

def build_week_card_html(week_num, week_start, week_end,
                         hypothesis, ga4p, hs, top, prev_rate) -> str:
    total_sess   = ga4p.get('total_sessions', 0)
    overall_rate = ga4p.get('overall_inquiry_rate', 0)
    top_slug     = top[0][0] if top else '—'
    top_rate     = top[0][1]['inquiry_rate'] if top else 0
    num_props    = len([d for d in ga4p.get('property_pages', {}).values()
                        if d['sessions'] >= MIN_SESSIONS])

    funnel    = hs.get('funnel_summary', {})
    new_deals = hs.get('new_deals_period', 0)
    converted = hs.get('converted_period', 0)
    book_rate = round(converted / max(new_deals, 1) * 100, 1)

    is_winner    = overall_rate > prev_rate
    winner_html  = '<span class="week-winner-badge">Improved</span>' if is_winner else ''
    hyp_display  = hypothesis[:180] + ('...' if len(hypothesis) > 180 else '')
    rate_delta   = f'{overall_rate:.2f}% (prev: {prev_rate:.2f}%)' if prev_rate else f'{overall_rate:.2f}%'

    return f"""
            <!-- Week {week_num} Card (auto-generated {datetime.now().strftime('%Y-%m-%d')}) -->
            <div class="week-card" id="week-{week_num}">
              <div class="week-card-header">
                <div class="week-number-badge">Week {week_num}</div>
                <div class="week-date-range">{week_start} &rarr; {week_end}</div>
                {winner_html}
              </div>
              <div class="week-hypothesis">
                <div class="week-section-label">Hypothesis</div>
                <p>{hyp_display}</p>
              </div>
              <div class="week-metrics-grid">
                <div class="week-metric">
                  <div class="wm-val">{total_sess:,}</div>
                  <div class="wm-lbl">Prop. Sessions</div>
                </div>
                <div class="week-metric">
                  <div class="wm-val">{rate_delta}</div>
                  <div class="wm-lbl">Inquiry Rate</div>
                </div>
                <div class="week-metric">
                  <div class="wm-val">{top_slug}</div>
                  <div class="wm-lbl">Top Property</div>
                </div>
                <div class="week-metric">
                  <div class="wm-val">{book_rate}%</div>
                  <div class="wm-lbl">Booking Rate</div>
                </div>
              </div>
              <div class="week-delta">
                <div class="week-section-label">Focus This Week</div>
                <p>Top: <strong>{top_slug}</strong> ({top_rate:.2f}%) &mdash;
                   {num_props} properties tracked.</p>
              </div>
              <a href="week-{week_num}.html" target="_blank" class="week-view-btn">View Week {week_num} Briefing &rarr;</a>
            </div>"""


def inject_week_card_into_index(week_card_html: str, week_num: int):
    if not INDEX_PATH.exists():
        print(f"[warn] {INDEX_PATH} not found — skipping index.html update")
        return

    content = INDEX_PATH.read_text(encoding='utf-8')

    coming_soon = re.compile(
        r'\s*<!-- week-coming-soon-start -->.*?<!-- week-coming-soon-end -->',
        re.DOTALL
    )
    content = coming_soon.sub('', content)

    marker = '<!-- /weekly-evolution-cards -->'
    if marker in content:
        content = content.replace(marker, week_card_html + '\n          ' + marker)
        INDEX_PATH.write_text(content, encoding='utf-8')
        print(f"[index] Injected Week {week_num} card into property-pages/index.html")
    else:
        print(f"[warn] Marker '{marker}' not found in index.html — skipping card injection")


# ── Git helpers ───────────────────────────────────────────────────────────────

def git_run(args, cwd=REPO_ROOT) -> str:
    result = subprocess.run(['git'] + args, cwd=str(cwd), capture_output=True, text=True)
    if result.returncode != 0:
        print(f"[git] stderr: {result.stderr.strip()}")
    return result.stdout.strip()


def commit_and_push(week_num: int, html_filename: str):
    files = [
        str(PROPERTY_DIR / html_filename),
        str(INDEX_PATH),
    ]
    if TRACKER_PATH.exists():
        files.append(str(TRACKER_PATH))

    for f in files:
        if Path(f).exists():
            git_run(['add', f])
            print(f"[git] Staged: {f}")

    msg = (
        f"Add Property CRO Week {week_num} briefing — automated weekly cycle\n\n"
        f"Generated {datetime.now().strftime('%Y-%m-%d')} by weekly_cycle_property.py.\n"
        f"Property tracker + index.html updated."
    )
    out = git_run(['commit', '-m', msg])
    if out:
        print(f"[git] Committed: {out[:80]}")

    push_out = git_run(['push'])
    print(f"[git] {'Push output: ' + push_out[:80] if push_out else 'Pushed to remote.'}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("\n" + "=" * 65)
    print("  THE GOOD LIFE BAHAMAS — WEEKLY PROPERTY PAGE CRO CYCLE")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 65 + "\n")

    PROPERTY_DIR.mkdir(exist_ok=True)

    # ── 1. Pull data
    print("[step 1/6] Pulling GA4 property page data (last 7 days)...")
    ga4p = pull_property_ga4_data(days=7)
    print(f"          {ga4p['total_sessions']:,} sessions across {len(ga4p['property_pages'])} property pages")
    print(f"          {ga4p['total_inquiries']} inquiries — overall rate {ga4p['overall_inquiry_rate']:.2f}%")

    print("[step 1/6] Pulling HubSpot pipeline data...")
    hs = pull_hubspot_data(days=7)
    funnel = hs.get('funnel_summary', {})
    print(f"          HubSpot: {funnel.get('total_deals', 0)} pipeline deals, "
          f"{hs.get('new_deals_period', 0)} new this period")

    print("[step 1/6] Pulling Clarity behavioural data...")
    clarity = pull_clarity_data(days=7)
    if clarity.get('available'):
        ov = clarity.get('overall', {})
        print(f"          Clarity: {ov.get('sessions', 0):,} sessions | "
              f"rage={ov.get('rage_click_pct', 0)}% | scroll={ov.get('avg_scroll_depth', 0)}%")
    else:
        print(f"          Clarity: not configured — add CLARITY_PROJECT_ID + CLARITY_API_KEY to .env")

    # ── 2. Rank properties + build hypothesis
    print("[step 2/6] Ranking properties and deriving hypothesis...")
    top, bottom = rank_properties(ga4p)
    week_num    = detect_week_number()
    week_start, week_end = get_week_date_range(week_num)
    print(f"          Week {week_num} ({week_start} → {week_end})")

    if top:
        print(f"          Top property: {top[0][0]} ({top[0][1]['inquiry_rate']:.2f}%)")
    if bottom:
        print(f"          Needs work:   {bottom[0][0]} ({bottom[0][1]['inquiry_rate']:.2f}%)")

    hypothesis = pick_property_hypothesis(ga4p, hs, top, bottom)
    print(f"          Hypothesis: {hypothesis[:100]}...")

    # Previous week's inquiry rate for comparison
    prev_rate = 0.0

    # ── 3. Generate briefing HTML
    print("[step 3/6] Generating Week {week_num} briefing HTML...")
    html_filename = f'week-{week_num}.html'
    html_content  = generate_briefing_html(
        week_num=week_num, week_start=week_start, week_end=week_end,
        hypothesis=hypothesis, ga4p=ga4p, hs=hs, top=top, bottom=bottom,
    )
    out_path = PROPERTY_DIR / html_filename
    out_path.write_text(html_content, encoding='utf-8')
    print(f"          Written: {out_path}")

    # ── 4. Update Excel tracker
    print("[step 4/6] Updating GoodLife_Property_CRO_Tracker.xlsx...")
    append_tracker_row(
        week_num=week_num, week_start=week_start, week_end=week_end,
        hypothesis=hypothesis, ga4p=ga4p, hs=hs, clarity=clarity,
        top=top, html_filename=html_filename,
    )

    # ── 5. Update property-pages/index.html
    print("[step 5/6] Updating property-pages/index.html...")
    week_card = build_week_card_html(
        week_num=week_num, week_start=week_start, week_end=week_end,
        hypothesis=hypothesis, ga4p=ga4p, hs=hs, top=top, prev_rate=prev_rate,
    )
    inject_week_card_into_index(week_card, week_num)

    # ── 6. Commit + push
    print("[step 6/6] Committing and pushing...")
    try:
        commit_and_push(week_num, html_filename)
    except Exception as e:
        print(f"[warn] Git step failed: {e}")
        print("       Files written — commit manually if needed.")

    # ── Summary
    print("\n" + "=" * 65)
    print(f"  DONE — Week {week_num} property briefing generated")
    print(f"  Briefing: property-pages/{html_filename}")
    print(f"  Tracker:  GoodLife_Property_CRO_Tracker.xlsx")
    print(f"  Overall inquiry rate this week: {ga4p['overall_inquiry_rate']:.2f}%")
    if top:
        print(f"  Top property: {top[0][0]} ({top[0][1]['inquiry_rate']:.2f}%)")
    print("=" * 65 + "\n")


if __name__ == '__main__':
    main()

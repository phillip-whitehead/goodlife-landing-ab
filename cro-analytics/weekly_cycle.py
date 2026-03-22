#!/usr/bin/env python3
"""
Good Life Bahamas — Weekly CRO Orchestration Script
=====================================================
Runs every week to:
  1. Pull GA4 + HubSpot data via cro_report.py
  2. Identify the top-performing version and top insight
  3. Generate a new week-N.html landing page based on the winner
  4. Append a metrics row to GoodLife_CRO_Tracker.xlsx
  5. Update versions/index.html to add the new week card
  6. Commit and push the new version

Usage:
  python3 weekly_cycle.py

Required env vars (or .env file):
  GA4_KEY_FILE     — path to service account JSON
  GA4_PROPERTY_ID  — GA4 property ID (e.g. 375125067)
  HUBSPOT_TOKEN    — HubSpot private app access token
"""

import os
import sys
import re
import subprocess
import shutil
from datetime import datetime, timedelta
from pathlib import Path

# ── Load .env if present ──────────────────────────────────────────────────────
try:
    from dotenv import load_dotenv
    _env_path = Path(__file__).parent / '.env'
    if _env_path.exists():
        load_dotenv(_env_path)
        print(f"[env] Loaded .env from {_env_path}")
except ImportError:
    pass

# ── Paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR    = Path(__file__).parent                          # cro-analytics/
REPO_ROOT     = SCRIPT_DIR.parent                              # repo root
VERSIONS_DIR  = REPO_ROOT / 'versions'                        # versions/
TRACKER_PATH  = REPO_ROOT / 'GoodLife_CRO_Tracker.xlsx'       # Excel tracker
INDEX_PATH    = VERSIONS_DIR / 'index.html'                   # versions/index.html

# ── Import from cro_report ────────────────────────────────────────────────────
sys.path.insert(0, str(SCRIPT_DIR))
from cro_report import (
    pull_ga4_data,
    pull_hubspot_data,
    pull_clarity_data,
    generate_insights,
    clarity_insights,
    AB_VERSIONS,
)


# ── Week detection ────────────────────────────────────────────────────────────

def detect_week_number() -> int:
    """Count existing week-N.html files in versions/ to get next week number."""
    existing = list(VERSIONS_DIR.glob('week-*.html'))
    week_nums = []
    for f in existing:
        m = re.match(r'^week-(\d+)\.html$', f.name)
        if m:
            week_nums.append(int(m.group(1)))
    if not week_nums:
        return 1
    return max(week_nums) + 1


def get_week_date_range(week_number: int) -> tuple[str, str]:
    """Return (start_date_str, end_date_str) for the given week number.
    Week 1 ends today; each prior week counts back 7 days per week."""
    end = datetime.now()
    start = end - timedelta(days=7)
    # Shift back by (week_number - 1) additional weeks for historical weeks
    offset = timedelta(days=7 * (week_number - 1))
    return (start - offset).strftime('%Y-%m-%d'), (end - offset).strftime('%Y-%m-%d')


# ── Version selection ─────────────────────────────────────────────────────────

def pick_base_version(ga4: dict, hs: dict) -> tuple[str, str, float]:
    """
    Returns (version_key, version_label, conversion_rate).
    Picks the version with highest HubSpot-contacts-per-session rate.
    Falls back to 'v5-best-combined' if no version traffic data exists.
    """
    version_pages   = ga4.get('version_pages', {})
    contacts_by_v   = hs.get('contacts_by_version', {})

    best_key   = None
    best_conv  = -1.0
    best_label = ''

    for vk, vdata in version_pages.items():
        sessions = vdata.get('sessions', 0)
        if sessions < 10:          # not enough data
            continue
        contacts = contacts_by_v.get(vk, 0)
        conv = contacts / sessions
        if conv > best_conv:
            best_conv  = conv
            best_key   = vk
            best_label = AB_VERSIONS.get(vk, vk)

    # Also check week-N versions (stored in version_pages as week-N)
    for vk, vdata in version_pages.items():
        if not vk.startswith('week-'):
            continue
        sessions = vdata.get('sessions', 0)
        if sessions < 10:
            continue
        contacts = contacts_by_v.get(vk, 0)
        conv = contacts / sessions
        if conv > best_conv:
            best_conv  = conv
            best_key   = vk
            best_label = f'Week {vk.split("-")[1]} iteration'

    if best_key is None:
        # Default to v5-best-combined as the canonical starting point
        best_key   = 'v5-best-combined'
        best_label = AB_VERSIONS.get('v5-best-combined', 'V5 — Best Combined')
        best_conv  = 0.0

    return best_key, best_label, round(best_conv * 100, 3)


def get_base_html_path(version_key: str) -> Path:
    """Resolve the HTML file path for a version key."""
    # Week versions: week-3.html
    if version_key.startswith('week-'):
        p = VERSIONS_DIR / f'{version_key}.html'
        if p.exists():
            return p

    # Named versions: v5-best-combined.html
    p = VERSIONS_DIR / f'{version_key}.html'
    if p.exists():
        return p

    # Final fallback
    fallback = VERSIONS_DIR / 'v5-best-combined.html'
    print(f"[warn] Could not find {version_key}.html — using v5-best-combined.html")
    return fallback


# ── Hypothesis generation ─────────────────────────────────────────────────────

def pick_hypothesis(insights: list[str], recs: list[str], ga4: dict, hs: dict) -> str:
    """
    Derives a single testable hypothesis string from the top recommendation
    and the overall data picture.
    """
    # Pull top-level numbers for context
    events         = ga4.get('events', {})
    form_starts    = events.get('form_start', {}).get('count', 0)
    form_submits   = (events.get('form_submit', {}).get('count', 0) +
                      events.get('inquire_form_submit', {}).get('count', 0))
    total_sessions = events.get('session_start', {}).get('count', 1)
    by_device      = {d['device']: d for d in ga4.get('by_device', [])}
    mobile         = by_device.get('mobile', {})
    mob_pct        = (mobile.get('sessions', 0) / max(total_sessions, 1)) * 100
    funnel         = hs.get('funnel_summary', {})
    close_lost_pct = funnel.get('close_lost_pct', 0)

    # Decision tree — pick the most pressing lever
    if form_starts > 0 and form_submits / max(form_starts, 1) < 0.10:
        completion = round(form_submits / form_starts * 100, 1)
        return (
            f"Form completion is critically low ({completion}% of form-starters convert). "
            f"Hypothesis: Reducing the form to a single email field with inline social proof "
            f"('Join {form_starts} travellers') and moving it above the search bar will lift "
            f"inquiry rate by 20%+."
        )

    if mob_pct > 65:
        mob_bounce = round(mobile.get('bounce_rate', 0) * 100, 1)
        return (
            f"{mob_pct:.0f}% of traffic is mobile (bounce {mob_bounce}%). "
            f"Hypothesis: A mobile-first hero with the CTA as a sticky bottom button "
            f"(no scroll required) and a tap-to-call option will reduce mobile bounce by 15%+."
        )

    if close_lost_pct > 70:
        return (
            f"{close_lost_pct}% of deals end in Close Lost. "
            f"Hypothesis: Adding a 'Get a reply within 2 hours' urgency signal next to the "
            f"inquiry form — with a countdown timer and host photo — will increase form-to-booked "
            f"conversion by surfacing the speed-of-response trust signal."
        )

    if recs:
        # Use the first recommendation, truncated to a hypothesis-length sentence
        top_rec = recs[0]
        # Trim to ~200 chars max
        if len(top_rec) > 200:
            top_rec = top_rec[:197] + '...'
        return f"Top insight-driven hypothesis: {top_rec}"

    return (
        "Iterative refinement: Increase social proof density in the above-the-fold area "
        "by adding a live 'X guests booked this month' counter and a featured guest photo "
        "testimonial to improve first-impression trust signals."
    )


# ── HTML generation ───────────────────────────────────────────────────────────

COMMENT_BLOCK_TEMPLATE = """\
<!--
  ============================================================
  WEEK {week_num} — AUTOMATED CRO ITERATION
  ============================================================
  Generated : {date_str}
  Base version  : {base_label} ({base_key})
  Date range    : {week_start} → {week_end}

  HYPOTHESIS:
  {hypothesis}

  KEY METRICS THAT DROVE THIS ITERATION:
    Sessions (7d)       : {sessions:,}
    Bounce rate         : {bounce_pct:.1f}%
    Form starts         : {form_starts:,}
    Form completion     : {form_completion:.1f}%
    Inquiry submissions : {form_submits:,}
    New HubSpot deals   : {new_deals}
    Converted (period)  : {converted}
    Close lost rate     : {close_lost_pct:.1f}%
    Winning base conv % : {base_conv_pct:.3f}%
    Mobile traffic %    : {mobile_pct:.0f}%

  CHANGES FROM BASE:
    - Version cookie updated to "week-{week_num}"
    - Comment block updated with this week's hypothesis and metrics
    - (Additional UX changes should be layered in manually or via
      an LLM-assisted edit pass after running this script)
  ============================================================
-->
"""

COOKIE_SCRIPT_TEMPLATE = """\
<script>
(function() {{
  var VERSION = 'week-{week_num}';

  function getCookie(name) {{
    var m = document.cookie.match('(?:^|; )' + name + '=([^;]*)');
    return m ? decodeURIComponent(m[1]) : null;
  }}
  function setCookie(name, val, days) {{
    var d = new Date();
    d.setTime(d.getTime() + days * 864e5);
    document.cookie = name + '=' + encodeURIComponent(val) + ';expires=' + d.toUTCString() + ';path=/;SameSite=Lax';
  }}
  if (!getCookie('gl_version')) setCookie('gl_version', VERSION, 30);
  try {{ sessionStorage.setItem('gl_version', VERSION); }} catch(e) {{}}

  window.addEventListener('load', function() {{
    if (typeof gtag === 'function') {{
      gtag('event', 'ab_test', {{ ab_version: VERSION }});
    }}
  }});
}})();
</script>"""


def generate_week_html(
    week_num: int,
    base_key: str,
    base_label: str,
    base_conv_pct: float,
    hypothesis: str,
    ga4: dict,
    hs: dict,
    week_start: str,
    week_end: str,
) -> str:
    """
    Generate the HTML content for week-N.html.
    Starts from the base version HTML, injects the metadata comment block,
    and updates the version-tracking cookie.
    """
    base_path    = get_base_html_path(base_key)
    source_html  = base_path.read_text(encoding='utf-8')

    # ── Compute metrics for the comment block
    events         = ga4.get('events', {})
    by_device      = {d['device']: d for d in ga4.get('by_device', [])}
    total_sessions = sum(d['sessions'] for d in ga4.get('by_device', []))
    mobile         = by_device.get('mobile', {})
    mob_pct        = mobile.get('sessions', 0) / max(total_sessions, 1) * 100

    version_pages  = ga4.get('version_pages', {})
    base_vdata     = version_pages.get(base_key, {})
    bounce_pct     = base_vdata.get('bounce_rate', 0) * 100

    form_starts    = events.get('form_start', {}).get('count', 0)
    form_submits   = (events.get('form_submit', {}).get('count', 0) +
                      events.get('inquire_form_submit', {}).get('count', 0))
    form_completion = (form_submits / form_starts * 100) if form_starts > 0 else 0.0

    funnel         = hs.get('funnel_summary', {})
    new_deals      = hs.get('new_deals_period', 0)
    converted      = hs.get('converted_period', 0)
    close_lost_pct = funnel.get('close_lost_pct', 0)

    # Indent hypothesis for the comment block
    hyp_indented = ('\n  ' + '  ').join(hypothesis.split('\n'))

    comment = COMMENT_BLOCK_TEMPLATE.format(
        week_num      = week_num,
        date_str      = datetime.now().strftime('%Y-%m-%d'),
        base_label    = base_label,
        base_key      = base_key,
        week_start    = week_start,
        week_end      = week_end,
        hypothesis    = hyp_indented,
        sessions      = total_sessions,
        bounce_pct    = bounce_pct,
        form_starts   = form_starts,
        form_completion = form_completion,
        form_submits  = form_submits,
        new_deals     = new_deals,
        converted     = converted,
        close_lost_pct = close_lost_pct,
        base_conv_pct = base_conv_pct,
        mobile_pct    = mob_pct,
    )

    cookie_script = COOKIE_SCRIPT_TEMPLATE.format(week_num=week_num)

    # ── Strip any existing leading comment block (<!-- ... -->) before <!DOCTYPE
    # so we don't accumulate blocks across iterations.
    cleaned = re.sub(r'^\s*<!--.*?-->\s*', '', source_html, flags=re.DOTALL)

    # ── Inject the new comment block at the top
    new_html = comment + cleaned

    # ── Replace the existing version-tracking script block.
    # The pattern: (function() { ... var VERSION = '...' ... })();
    # We match from the inline cookie script open to its closing tag.
    cookie_pattern = re.compile(
        r'<script>\s*\(function\(\)\s*\{.*?var VERSION\s*=\s*[\'"][^\'"]*[\'"].*?\}\)\(\);\s*</script>',
        re.DOTALL
    )
    if cookie_pattern.search(new_html):
        new_html = cookie_pattern.sub(cookie_script, new_html, count=1)
    else:
        # Fallback: insert cookie script just before </body>
        new_html = new_html.replace('</body>', cookie_script + '\n</body>', 1)

    return new_html


# ── Excel tracker ─────────────────────────────────────────────────────────────

TRACKER_COLUMNS = [
    # ── Identity
    'Week',
    'Date Generated',
    'Week Start',
    'Week End',
    'Base Version',
    'Hypothesis (short)',
    # ── GA4
    'Total Sessions (7d)',
    'Mobile % of Sessions',
    'Avg Bounce Rate %',
    'Form Starts',
    'Form Submits',
    'Form Completion %',
    'Inquiry Rate % (submits/sessions)',
    # ── HubSpot
    'New HubSpot Deals (7d)',
    'Converted (Booked, 7d)',
    'Period Conversion %',
    'Close Lost % (all-time pipeline)',
    'Booked or Beyond (all-time)',
    # ── Clarity (auto-pulled)
    'Rage Click Rate %',
    'Dead Click Rate %',
    'Quick-Back Rate %',
    'Avg Scroll Depth %',
    'Excessive Scroll Rate %',
    'Avg Active Time (sec)',
    # ── Clarity (manual 15-min weekly review)
    'Clarity Notes (manual)',
    'Top Rage-Click Element (manual)',
    # ── Result
    'Best Version (by conv rate)',
    'Best Version Conv Rate %',
    'HTML File',
    'Clarity Dashboard Link',
]


def append_tracker_row(
    week_num: int,
    week_start: str,
    week_end: str,
    base_key: str,
    base_label: str,
    hypothesis: str,
    ga4: dict,
    hs: dict,
    clarity: dict,
    best_key: str,
    best_conv: float,
    html_filename: str,
):
    """Append a metrics row to GoodLife_CRO_Tracker.xlsx (create if missing)."""
    try:
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("[warn] openpyxl not installed — skipping Excel update. Run: pip install openpyxl")
        return

    # ── Compute metrics
    events         = ga4.get('events', {})
    by_device      = {d['device']: d for d in ga4.get('by_device', [])}
    total_sessions = sum(d['sessions'] for d in ga4.get('by_device', []))
    mobile         = by_device.get('mobile', {})
    mob_pct        = round(mobile.get('sessions', 0) / max(total_sessions, 1) * 100, 1)

    version_pages  = ga4.get('version_pages', {})
    all_bounces    = [v['bounce_rate'] for v in version_pages.values() if v.get('sessions', 0) > 0]
    avg_bounce     = round(sum(all_bounces) / len(all_bounces) * 100, 1) if all_bounces else 0.0

    form_starts    = events.get('form_start', {}).get('count', 0)
    form_submits   = (events.get('form_submit', {}).get('count', 0) +
                      events.get('inquire_form_submit', {}).get('count', 0))
    form_completion = round(form_submits / form_starts * 100, 1) if form_starts > 0 else 0.0
    inquiry_rate    = round(form_submits / max(total_sessions, 1) * 100, 3)

    funnel         = hs.get('funnel_summary', {})
    new_deals      = hs.get('new_deals_period', 0)
    converted      = hs.get('converted_period', 0)
    period_conv    = round(converted / new_deals * 100, 1) if new_deals > 0 else 0.0
    close_lost_pct = funnel.get('close_lost_pct', 0)
    booked_beyond  = funnel.get('booked_or_beyond', 0)

    # Clarity metrics
    c_overall      = clarity.get('overall', {}) if clarity.get('available') else {}
    c_rage         = c_overall.get('rage_click_pct',      '')
    c_dead         = c_overall.get('dead_click_pct',      '')
    c_qback        = c_overall.get('quick_back_pct',      '')
    c_scroll       = c_overall.get('avg_scroll_depth',    '')
    c_xscroll      = c_overall.get('excessive_scroll_pct','')
    c_active       = c_overall.get('avg_active_time',     '')
    c_link         = clarity.get('clarity_link', '')

    hyp_short = hypothesis[:120] + ('...' if len(hypothesis) > 120 else '')

    row_data = [
        # Identity
        week_num,
        datetime.now().strftime('%Y-%m-%d'),
        week_start,
        week_end,
        f'{base_label} ({base_key})',
        hyp_short,
        # GA4
        total_sessions,
        mob_pct,
        avg_bounce,
        form_starts,
        form_submits,
        form_completion,
        inquiry_rate,
        # HubSpot
        new_deals,
        converted,
        period_conv,
        close_lost_pct,
        booked_beyond,
        # Clarity (auto)
        c_rage,
        c_dead,
        c_qback,
        c_scroll,
        c_xscroll,
        c_active,
        # Clarity (manual — leave blank, fill in after weekly heatmap review)
        '',
        '',
        # Result
        AB_VERSIONS.get(best_key, best_key),
        best_conv,
        html_filename,
        c_link,
    ]

    # ── Load or create workbook
    if TRACKER_PATH.exists():
        wb = load_workbook(TRACKER_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Weekly CRO Tracker'

        # Header row styling
        header_fill = PatternFill(start_color='0D1B2A', end_color='0D1B2A', fill_type='solid')
        header_font = Font(bold=True, color='31B8C6', size=10)
        thin        = Side(style='thin', color='31B8C6')
        header_border = Border(bottom=Side(style='medium', color='31B8C6'))

        for col_idx, col_name in enumerate(TRACKER_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font   = header_font
            cell.fill   = header_fill
            cell.border = header_border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        # Column widths (approximate)
        col_widths = [8, 14, 12, 12, 30, 60, 16, 16, 14, 12, 12, 16, 20, 16, 16,
                      16, 20, 20, 30, 18, 28]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w

        # Freeze header
        ws.freeze_panes = 'A2'

    # ── Append data row
    next_row = ws.max_row + 1
    data_font = Font(size=9)
    alt_fill  = PatternFill(start_color='1A2B3C', end_color='1A2B3C', fill_type='solid')
    row_fill  = alt_fill if next_row % 2 == 0 else None

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=value)
        cell.font = data_font
        cell.alignment = Alignment(vertical='top', wrap_text=(col_idx == 6))
        if row_fill:
            cell.fill = row_fill

    wb.save(TRACKER_PATH)
    print(f"[tracker] Appended Week {week_num} row to {TRACKER_PATH}")


# ── versions/index.html update ────────────────────────────────────────────────

def build_week_card_html(
    week_num: int,
    week_start: str,
    week_end: str,
    hypothesis: str,
    ga4: dict,
    hs: dict,
    base_label: str,
    best_conv: float,
    prev_best_conv: float,
) -> str:
    """Generate the HTML snippet for one week card in the Weekly Evolution section."""
    events         = ga4.get('events', {})
    by_device      = {d['device']: d for d in ga4.get('by_device', [])}
    total_sessions = sum(d['sessions'] for d in ga4.get('by_device', []))
    mobile         = by_device.get('mobile', {})
    mob_bounce_raw = mobile.get('bounce_rate', 0)

    version_pages  = ga4.get('version_pages', {})
    all_bounces    = [v['bounce_rate'] for v in version_pages.values() if v.get('sessions', 0) > 0]
    avg_bounce     = (sum(all_bounces) / len(all_bounces) * 100) if all_bounces else 0.0

    form_starts    = events.get('form_start', {}).get('count', 0)
    form_submits   = (events.get('form_submit', {}).get('count', 0) +
                      events.get('inquire_form_submit', {}).get('count', 0))
    inquiry_rate   = round(form_submits / max(total_sessions, 1) * 100, 2)

    funnel         = hs.get('funnel_summary', {})
    new_deals      = hs.get('new_deals_period', 0)
    converted      = hs.get('converted_period', 0)
    booking_rate   = round(converted / max(new_deals, 1) * 100, 1)

    is_winner  = best_conv > prev_best_conv
    winner_html = '<span class="week-winner-badge">Winner</span>' if is_winner else ''

    hyp_display = hypothesis[:180] + ('...' if len(hypothesis) > 180 else '')

    what_changed = f'Iterated on <strong>{base_label}</strong>.'
    if is_winner:
        what_changed += f' Improved conversion from {prev_best_conv:.3f}% to {best_conv:.3f}%.'
    else:
        what_changed += f' Conversion: {best_conv:.3f}% (baseline: {prev_best_conv:.3f}%).'

    return f"""
            <!-- Week {week_num} Card (auto-generated {datetime.now().strftime('%Y-%m-%d')}) -->
            <div class="week-card" id="week-{week_num}">
              <div class="week-card-header">
                <div class="week-number-badge">Week {week_num}</div>
                <div class="week-date-range">{week_start} &rarr; {week_end}</div>
                {winner_html}
              </div>
              <div class="week-hypothesis">
                <div class="week-section-label">Hypothesis</div>
                <p>{hyp_display}</p>
              </div>
              <div class="week-metrics-grid">
                <div class="week-metric">
                  <div class="wm-val">{total_sessions:,}</div>
                  <div class="wm-lbl">Sessions</div>
                </div>
                <div class="week-metric">
                  <div class="wm-val">{avg_bounce:.1f}%</div>
                  <div class="wm-lbl">Avg Bounce</div>
                </div>
                <div class="week-metric">
                  <div class="wm-val">{inquiry_rate:.2f}%</div>
                  <div class="wm-lbl">Inquiry Rate</div>
                </div>
                <div class="week-metric">
                  <div class="wm-val">{booking_rate:.1f}%</div>
                  <div class="wm-lbl">Booking Rate</div>
                </div>
              </div>
              <div class="week-delta">
                <div class="week-section-label">What Changed</div>
                <p>{what_changed}</p>
              </div>
              <a href="week-{week_num}.html" target="_blank" class="week-view-btn">View Week {week_num} &rarr;</a>
            </div>"""


def inject_week_card_into_index(week_card_html: str, week_num: int):
    """
    Insert a new week card into the weekly-evolution-cards container
    in versions/index.html, replacing the 'coming soon' empty state if present.
    """
    if not INDEX_PATH.exists():
        print(f"[warn] {INDEX_PATH} not found — skipping index.html update")
        return

    content = INDEX_PATH.read_text(encoding='utf-8')

    # Remove the coming-soon placeholder if present
    coming_soon_pattern = re.compile(
        r'\s*<!-- week-coming-soon-start -->.*?<!-- week-coming-soon-end -->',
        re.DOTALL
    )
    content = coming_soon_pattern.sub('', content)

    # Inject the new card before the closing marker of the cards container
    marker = '<!-- /weekly-evolution-cards -->'
    if marker in content:
        content = content.replace(marker, week_card_html + '\n          ' + marker)
        INDEX_PATH.write_text(content, encoding='utf-8')
        print(f"[index] Injected Week {week_num} card into versions/index.html")
    else:
        print(f"[warn] Marker '{marker}' not found in index.html — skipping card injection")


# ── Git helpers ───────────────────────────────────────────────────────────────

def git_run(args: list[str], cwd: Path = REPO_ROOT) -> str:
    """Run a git command and return stdout."""
    result = subprocess.run(
        ['git'] + args,
        cwd=str(cwd),
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        print(f"[git] stderr: {result.stderr.strip()}")
    return result.stdout.strip()


def commit_and_push(week_num: int, html_filename: str):
    """Stage new/changed files, commit, and push."""
    files_to_stage = [
        str(VERSIONS_DIR / html_filename),
        str(INDEX_PATH),
    ]
    if TRACKER_PATH.exists():
        files_to_stage.append(str(TRACKER_PATH))

    for f in files_to_stage:
        if Path(f).exists():
            git_run(['add', f])
            print(f"[git] Staged: {f}")

    # Also stage any new report HTML
    report_glob = list((SCRIPT_DIR / 'reports').glob('cro_report_*.html'))
    if report_glob:
        latest_report = sorted(report_glob)[-1]
        git_run(['add', str(latest_report)])
        print(f"[git] Staged report: {latest_report}")

    commit_msg = (
        f"Add Week {week_num} CRO iteration — automated weekly cycle\n\n"
        f"Generated {datetime.now().strftime('%Y-%m-%d')} by weekly_cycle.py.\n"
        f"Base: auto-selected winner. Tracker + index.html updated."
    )
    out = git_run(['commit', '-m', commit_msg])
    if out:
        print(f"[git] Committed: {out[:80]}")

    push_out = git_run(['push'])
    if push_out:
        print(f"[git] Push output: {push_out[:120]}")
    else:
        print("[git] Pushed to remote.")


# ── Main orchestration ────────────────────────────────────────────────────────

def main():
    print("\n" + "=" * 65)
    print("  THE GOOD LIFE BAHAMAS — WEEKLY CRO CYCLE")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 65 + "\n")

    # ── 1. Pull analytics data
    print("[step 1/6] Pulling GA4 data (last 7 days)...")
    ga4 = pull_ga4_data(days=7)
    total_sessions = sum(d['sessions'] for d in ga4.get('by_device', []))
    print(f"          GA4: {total_sessions:,} sessions, "
          f"{len(ga4.get('version_pages', {}))} version pages tracked")

    print("[step 1/6] Pulling HubSpot pipeline data...")
    hs = pull_hubspot_data(days=7)
    funnel = hs.get('funnel_summary', {})
    print(f"          HubSpot: {funnel.get('total_deals', 0)} pipeline deals, "
          f"{hs.get('new_deals_period', 0)} new this period")

    print("[step 1/6] Pulling Clarity behavioural data...")
    clarity = pull_clarity_data(days=7)
    if clarity.get('available'):
        ov = clarity.get('overall', {})
        print(f"          Clarity: {ov.get('sessions', 0):,} sessions | "
              f"rage={ov.get('rage_click_pct', 0)}% | "
              f"dead={ov.get('dead_click_pct', 0)}% | "
              f"scroll depth={ov.get('avg_scroll_depth', 0)}%")
    else:
        print(f"          Clarity: not configured — add CLARITY_PROJECT_ID + CLARITY_API_KEY to .env")

    # ── 2. Generate insights + save HTML report
    print("[step 2/6] Generating insights...")
    insights, recs = generate_insights(ga4, hs)
    c_insights, c_recs = clarity_insights(clarity)
    insights += c_insights
    recs     += c_recs
    print(f"          {len(insights)} insights, {len(recs)} recommendations")

    reports_dir = SCRIPT_DIR / 'reports'
    reports_dir.mkdir(exist_ok=True)
    from cro_report import save_html_report
    report_fname = reports_dir / f"cro_report_{datetime.now().strftime('%Y%m%d')}.html"
    save_html_report(ga4, hs, insights, recs, days=7, output_path=str(report_fname))

    # ── 3. Detect week number + select base version
    print("[step 3/6] Detecting week number and selecting base version...")
    week_num = detect_week_number()
    week_start, week_end = get_week_date_range(week_num)
    print(f"          This is Week {week_num} ({week_start} → {week_end})")

    base_key, base_label, base_conv = pick_base_version(ga4, hs)
    print(f"          Base version: {base_label} ({base_key})  conv={base_conv:.3f}%")

    # ── 4. Build hypothesis
    print("[step 4/6] Deriving hypothesis from top insight...")
    hypothesis = pick_hypothesis(insights, recs, ga4, hs)
    print(f"          Hypothesis: {hypothesis[:100]}...")

    # Get previous week's conversion for winner badge comparison
    prev_week_num = week_num - 1
    prev_conv = 0.0
    if prev_week_num >= 1:
        prev_version_key = f'week-{prev_week_num}'
        prev_contacts = hs.get('contacts_by_version', {}).get(prev_version_key, 0)
        prev_sessions = ga4.get('version_pages', {}).get(prev_version_key, {}).get('sessions', 1)
        prev_conv = prev_contacts / prev_sessions * 100

    # ── 5. Generate HTML
    print("[step 5/6] Generating week HTML...")
    html_filename = f'week-{week_num}.html'
    html_content  = generate_week_html(
        week_num      = week_num,
        base_key      = base_key,
        base_label    = base_label,
        base_conv_pct = base_conv,
        hypothesis    = hypothesis,
        ga4           = ga4,
        hs            = hs,
        week_start    = week_start,
        week_end      = week_end,
    )
    output_path = VERSIONS_DIR / html_filename
    output_path.write_text(html_content, encoding='utf-8')
    print(f"          Written: {output_path}")

    # ── 5b. Update Excel tracker
    print("[step 5/6] Updating GoodLife_CRO_Tracker.xlsx...")
    append_tracker_row(
        week_num   = week_num,
        week_start = week_start,
        week_end   = week_end,
        base_key   = base_key,
        base_label = base_label,
        hypothesis = hypothesis,
        ga4        = ga4,
        hs         = hs,
        clarity    = clarity,
        best_key   = base_key,
        best_conv  = base_conv,
        html_filename = html_filename,
    )

    # ── 5c. Update versions/index.html
    print("[step 5/6] Updating versions/index.html...")
    week_card = build_week_card_html(
        week_num      = week_num,
        week_start    = week_start,
        week_end      = week_end,
        hypothesis    = hypothesis,
        ga4           = ga4,
        hs            = hs,
        base_label    = base_label,
        best_conv     = base_conv,
        prev_best_conv = prev_conv,
    )
    inject_week_card_into_index(week_card, week_num)

    # ── 6. Commit and push
    print("[step 6/6] Committing and pushing...")
    try:
        commit_and_push(week_num, html_filename)
    except Exception as e:
        print(f"[warn] Git step failed: {e}")
        print("       Files are written — commit manually if needed.")

    # ── Summary
    print("\n" + "=" * 65)
    print(f"  WEEKLY CYCLE COMPLETE — Week {week_num}")
    print("=" * 65)
    print(f"  New page   : versions/{html_filename}")
    print(f"  Tracker    : {TRACKER_PATH.name}")
    print(f"  Report     : {report_fname.name}")
    print(f"  Base       : {base_label}")
    print(f"  Sessions   : {total_sessions:,}")
    form_starts  = ga4.get('events', {}).get('form_start', {}).get('count', 0)
    form_submits = (ga4.get('events', {}).get('form_submit', {}).get('count', 0) +
                    ga4.get('events', {}).get('inquire_form_submit', {}).get('count', 0))
    print(f"  Form funnel: {form_starts} starts → {form_submits} submits")
    print(f"  New deals  : {hs.get('new_deals_period', 0)}")
    print(f"  Hypothesis : {hypothesis[:90]}...")
    print("=" * 65 + "\n")


if __name__ == '__main__':
    main()
